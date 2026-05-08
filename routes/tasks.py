from flask import Blueprint, request, jsonify, render_template, flash, url_for, redirect, abort, current_app, send_file
from io import BytesIO
import os
import uuid
from flask_jwt_extended import (
    jwt_required,
    get_jwt_identity,
    get_jwt
)
from datetime import datetime, date
from sqlalchemy import desc, exists, asc
from extensions import db, allowed_file
from models.task import Task, TaskUserAssignment
from models.user import User
from models.role import Role
from models.group import Group, UserGroup
from models.attachment import Attachment
from models.announcement import Announcement, AnnouncementView
from models.meeting import MeetingTask, DepartmentMeeting
from decorators.roles import roles_required

tasks_bp = Blueprint("tasks", __name__, url_prefix="/tasks")


@tasks_bp.route("/", methods=["GET"])
@jwt_required()
def list_tasks():
    user_id = get_jwt_identity()
    role = Role.query.filter_by(id=get_jwt()["role"]).first()
    users = []
    groups = []
    groups_members = {}
    page = request.args.get("page", 1, type=int)
    per_page = 10

    user_assignments = {
        int(a.task_id): a
        for a in TaskUserAssignment.query.filter_by(user_id=user_id).all()
    }

    task_ids_by_assignment = [int(r[0]) for r in db.session.query(TaskUserAssignment.task_id).filter(TaskUserAssignment.user_id == user_id).all()]
    assigned_task_ids = set(task_ids_by_assignment)

    is_leader = role.name in ("Документовед", "Руководитель")

    if is_leader:
        users = User.query.filter(User.dismissal_date.is_(None)).all()
        groups = Group.query.all()
        for group in groups:
            members = (
                db.session.query(User)
                .join(UserGroup, UserGroup.user_id == User.id)
                .filter(UserGroup.group_id == group.id, User.dismissal_date.is_(None))
                .all()
            )
            groups_members[group.id] = [{"id": m.id, "name": m.name} for m in members]

    if is_leader:
        tasks = Task.query.order_by(asc(Task.deadline_at)).all()
    else:
        tasks = (
            Task.query
            .filter(Task.id.in_(task_ids_by_assignment))
            .order_by(asc(Task.deadline_at))
            .all()
        )

    all_assignments = {
        int(a.task_id): a
        for a in TaskUserAssignment.query.all()
    }

    review_task_ids = set()
    for tid, a in all_assignments.items():
        if a.status == 'на проверке':
            review_task_ids.add(tid)

    def task_sort_key(t):
        if is_leader:
            has_review = int(t.id) in review_task_ids
            has_completed = any(a.approved for aid, a in all_assignments.items() if aid == int(t.id))
            has_incomplete = any(not a.approved for aid, a in all_assignments.items() if aid == int(t.id))
            if has_review:
                return 0
            if not has_incomplete:
                return 1
            return 2
        asgn = user_assignments.get(int(t.id))
        if not asgn:
            return 0
        if asgn.approved:
            return 2
        if asgn.status == 'на проверке':
            return 1
        return 0

    tasks = sorted(tasks, key=task_sort_key)

    total = len(tasks)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_tasks = tasks[start:end]

    today_date = date.today()
    announcements = Announcement.query.filter(Announcement.is_deleted == False, Announcement.deadline >= today_date).order_by(Announcement.created_at.desc()).limit(15).all()
    viewed_ids = {v.announcement_id for v in AnnouncementView.query.filter_by(user_id=user_id).all()}

    return render_template("tasks/list.html", tasks=paginated_tasks, role=role, users=users, groups=groups, groups_members=groups_members, user_assignments=user_assignments, assigned_task_ids=assigned_task_ids, user_id=user_id, today=today_date, page=page, total=total, per_page=per_page, announcements=announcements, viewed_ids=viewed_ids, review_task_ids=review_task_ids, is_leader=is_leader)

@tasks_bp.route("/", methods=["POST"])
@jwt_required()
@roles_required("Руководитель", "Документовед")
def create_task():
    if request.method == 'POST':
        user_id = get_jwt_identity()
        role = Role.query.filter_by(id=get_jwt()["role"]).first()
        title = request.form['title']
        description = request.form['description']
        priority = request.form['priority']
        deadline_at = request.form['deadline_at']
        no_review = request.form.get('no_review') == 'on'
        assignees = request.form.getlist('assignees')
        assignees = [a for a in assignees if a.strip()]

        if not assignees:
            flash('Укажите исполнителей для задачи', 'danger')
            return redirect(url_for('tasks.list_tasks'))

        new_task = Task(
            title=title,
            description=description,
            priority=priority,
            deadline_at=deadline_at,
            no_review=no_review,
            creator_id=user_id
        )

        db.session.add(new_task)
        db.session.commit()

        files = request.files.getlist('files')
        if files and files[0].filename:
            for file in files:
                if file.filename:
                    if not allowed_file(file.filename, current_app.config):
                        flash(f'Недопустимый формат файла: {file.filename}', 'danger')
                        return redirect(url_for('tasks.task_details', task_id=new_task.id))

                    original_name = file.filename
                    ext = os.path.splitext(original_name)[1].lower()
                    safe_filename = str(uuid.uuid4()) + ext
                    upload_dir = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'attachments')
                    os.makedirs(upload_dir, exist_ok=True)
                    file_path = os.path.join(upload_dir, safe_filename)
                    file.save(file_path)
                    file.seek(0, 2)
                    file_size = file.tell()
                    attachment = Attachment(
                        task_id=new_task.id,
                        file_name=original_name,
                        file_path=file_path,
                        mime_type=file.content_type or 'application/octet-stream',
                        size=file_size
                    )
                    db.session.add(attachment)

        for assignee_id in assignees:
            task_assignment = TaskUserAssignment(task_id=new_task.id, user_id=assignee_id)
            db.session.add(task_assignment)

        db.session.commit()

        flash('Задача успешно создана!', 'success')
        return redirect(url_for('tasks.list_tasks'))

@tasks_bp.route("/filter", methods=["GET"])
@jwt_required()
def filter_tasks():
    user_id = get_jwt_identity()
    role = get_jwt()["role"]

    priority = request.args.get("priority")
    status = request.args.get("status")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    query = Task.query

    if role == 3:
        task_ids_by_assignment = [r[0] for r in db.session.query(TaskUserAssignment.task_id).filter(TaskUserAssignment.user_id == user_id).all()]
        query = query.filter(Task.id.in_(task_ids_by_assignment))

    if priority:
        query = query.filter(Task.priority == priority)

    if date_from:
        query = query.filter(Task.deadline_at >= datetime.fromisoformat(date_from))

    if date_to:
        query = query.filter(Task.deadline_at <= datetime.fromisoformat(date_to))

    tasks = query.all()

    if status:
        tasks = [
            t for t in tasks
            if any(
                s.status == status and s.user_id == user_id
                for s in t.statuses
            )
        ]

    return jsonify([task.to_dict() for task in tasks])


@tasks_bp.route("/calendar", methods=["GET"])
@jwt_required()
def calendar():
    user_id = get_jwt_identity()
    role = Role.query.filter_by(id=get_jwt()["role"]).first()
    is_leader = role.name in ("Документовед", "Руководитель")
    tasks = []

    user_assignments = {
        a.task_id: a
        for a in TaskUserAssignment.query.filter_by(user_id=user_id).all()
    }

    task_ids_by_assignment = [r[0] for r in db.session.query(TaskUserAssignment.task_id).filter(TaskUserAssignment.user_id == user_id).all()]

    if is_leader:
        tasks = Task.query.order_by(desc(Task.deadline_at)).all()
    else:
        tasks = (
            Task.query
            .filter(Task.id.in_(task_ids_by_assignment))
            .order_by(desc(Task.deadline_at))
            .all()
        )

    events = []
    today = date.today()
    for task in tasks:
        assignment = user_assignments.get(task.id)
        status = assignment.status if assignment else None
        is_overdue = status not in ('завершена', 'на проверке') and task.deadline_at < today
        color = None
        if is_overdue:
            color = 'darkred'
        elif task.priority == 'high':
            color = 'red'
        elif task.priority == 'medium':
            color = 'orange'
        else:
            color = 'light-blue'
        title = task.title
        if is_overdue:
            title = f'⚠ {task.title}'
        events.append({
            'title': title,
            'end': task.deadline_at.strftime('%Y-%m-%d'),
            'description': task.description,
            'color': color,
            'url': url_for('tasks.task_details', task_id=task.id),
            'status': status
        })
    tasks_by_deadline = {}
    today = date.today()
    for task in tasks:
        deadline_str = task.deadline_at.strftime('%Y-%m-%d')
        if deadline_str not in tasks_by_deadline:
            tasks_by_deadline[deadline_str] = []
        assignment = user_assignments.get(task.id)
        status = assignment.status if assignment else None
        tasks_by_deadline[deadline_str].append({
            'id': task.id,
            'title': task.title,
            'status': status,
            'priority': task.priority,
            'is_overdue': status is not None and status not in ('завершена', 'на проверке') and task.deadline_at < today
        })

    days_all_completed = {}
    today_date = date.today()
    days_has_unassigned = {}
    for day, day_tasks in tasks_by_deadline.items():
        assigned = [t for t in day_tasks if t['status'] is not None]
        days_all_completed[day] = len(assigned) > 0 and all(t['status'] in ('завершена', 'на проверке') for t in assigned)
        days_has_unassigned[day] = any(t['status'] is None for t in day_tasks)

    announcements = Announcement.query.filter(Announcement.is_deleted == False, Announcement.deadline >= today_date).order_by(Announcement.created_at.desc()).limit(15).all()
    viewed_ids = {v.announcement_id for v in AnnouncementView.query.filter_by(user_id=user_id).all()}

    return render_template("tasks/calendar.html", tasks=events, tasks_by_deadline=tasks_by_deadline, days_all_completed=days_all_completed, days_has_unassigned=days_has_unassigned, role=role, announcements=announcements, viewed_ids=viewed_ids, is_leader=is_leader)

@tasks_bp.route("/<int:task_id>", methods=["DELETE"])
@jwt_required()
@roles_required("Руководитель")
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)

    db.session.delete(task)
    db.session.commit()

    return jsonify({"msg": "Task deleted"})

@tasks_bp.route('/get/<int:task_id>', methods=['GET'])
@tasks_bp.route('/<int:task_id>', methods=['GET'])
@jwt_required()
def task_details(task_id):
    user_id = get_jwt_identity()
    role = Role.query.filter_by(id=get_jwt()["role"]).first()
    page = request.args.get('page', 1, type=int)
    from_page = request.args.get('from', None)
    from_meeting = request.args.get('from_meeting', None, type=int)

    if role.name == 'Сотрудник':
        has_assignment = TaskUserAssignment.query.filter_by(task_id=task_id, user_id=user_id).first()
        if not has_assignment:
            return abort(404)
        task = Task.query.filter_by(id=task_id).first_or_404()
    else:
        task = Task.query.filter_by(id=task_id).first_or_404()

    assignees = []
    is_assigned = False
    user_assignment = TaskUserAssignment.query.filter_by(task_id=task_id, user_id=user_id).first()
    if role.name != 'Сотрудник':
        assignees = TaskUserAssignment.query.filter_by(task_id=task_id).all()
        status_order = {'завершена': 0, 'на проверке': 1, 'в работе': 2, 'не начата': 3}
        assignees.sort(key=lambda a: (status_order.get(a.status, 4), a.user.name))

    if user_assignment:
        is_assigned = True

    if role.name == 'Сотрудник':
        assignees = [user_assignment] if user_assignment else []

    mt = MeetingTask.query.filter_by(task_id=task_id).first()
    meeting_id = None
    meeting_title = None
    if mt:
        m = DepartmentMeeting.query.get(mt.meeting_id)
        if m:
            meeting_id = m.id
            meeting_title = m.title

    return render_template('tasks/details.html', task=task, assignees=assignees, today=date.today(), user_id=user_id, user_role=role, is_assigned=is_assigned, user_assignment=user_assignment, page=page, from_page=from_page, meeting_id=meeting_id, meeting_title=meeting_title, from_meeting=from_meeting)

@tasks_bp.route('/<int:task_id>/report', methods=['GET'])
@jwt_required()
@roles_required("Руководитель", "Документовед")
def generate_task_report(task_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    task = Task.query.get_or_404(task_id)
    assignees = TaskUserAssignment.query.filter_by(task_id=task_id).all()

    completed = []
    in_progress = []
    on_review = []

    for a in assignees:
        entry = {"user_name": a.user.name, "status": a.status}
        if a.status == 'завершена':
            completed.append(entry)
        elif a.status == 'на проверке':
            on_review.append(entry)
        else:
            in_progress.append(entry)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по задаче"

    header_font = Font(name='Calibri', size=16, bold=True)
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11)
    section_font = Font(name='Calibri', size=13, bold=True)
    header_row_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=11)

    green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    light_bg = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=f"Отчёт по задаче: {task.title}")
    cell.font = header_font

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"Приоритет: {task.priority}").font = value_font
    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"Дедлайн: {task.deadline_at.strftime('%d.%m.%Y')}").font = value_font
    if task.description:
        row = 4
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=f"Описание: {task.description}").font = value_font

    row += 1
    total = len(assignees)
    done_count = len(completed)
    pct = round(done_count / total * 100) if total > 0 else 0
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"Прогресс: {done_count}/{total} ({pct}%)").font = Font(name='Calibri', size=12, bold=True, color='006100')

    row += 2

    def write_section(title, items, fill_color):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        row += 1

        for col_idx, col_name in enumerate(['№', 'Исполнитель', 'Статус'], 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_row_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        for i, entry in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i).font = cell_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=entry["user_name"]).font = cell_font
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=entry["status"]).font = cell_font
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).border = thin_border

            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = fill_color
            row += 1
        row += 1

    if completed:
        write_section(f"Выполнено — {len(completed)}", completed, green_fill)
    if on_review:
        write_section(f"На проверке — {len(on_review)}", on_review, yellow_fill)
    if in_progress:
        write_section(f"Не выполнено — {len(in_progress)}", in_progress, red_fill)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'report_{task.id}.xlsx')